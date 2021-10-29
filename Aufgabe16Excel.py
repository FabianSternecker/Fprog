import openpyxl
import math
wb = openpyxl.Workbook()
ws = wb.active
ws.title="MeineErsteTabelle"
x=-3.1
i=1
ws["A"+str(i)]="x"
ws["B"+str(i)]="f(x)"
i+=1
while(x<3.2):
    ws["A"+str(i)]= x
    ws["B" + str(i)] = x * math.sin(x)
    x += 0.1
    i += 1
wb.save("timeToExcel.xlsx")